from __future__ import annotations

from dataclasses import asdict
from datetime import date
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

from excel_assistant.models import AppSettings, MonitoredEntry
from excel_assistant.utils import (
    business_days_between,
    col_to_index,
    dedupe_rows,
    index_to_col,
    normalize_col,
    parse_list_csv,
    to_date,
)


def classify_status(days: int, settings: AppSettings) -> str:
    t = settings.thresholds
    if days <= t.good_max:
        return "good"
    if days <= t.soft_max:
        return "soft"
    if days <= t.medium_max:
        return "medium"
    if days <= t.hard_max:
        return "hard"
    if days >= t.due_at:
        return "due"
    return "hard"


class ExcelMonitor:
    def scan(self, settings: AppSettings) -> list[MonitoredEntry]:
        workbook_path = Path(settings.excel_file_path)
        if not workbook_path.exists():
            return []

        wb = load_workbook(workbook_path, read_only=True, data_only=True, keep_links=False)
        ws = wb[settings.watch.sheet_name] if settings.watch.sheet_name in wb.sheetnames else wb.active

        entries: list[MonitoredEntry] = []
        today = date.today()

        for row, col_idx in self._iter_targets(ws, settings):
            cell_ref = f"{index_to_col(col_idx)}{row}"
            value = ws.cell(row=row, column=col_idx).value
            entry_date = to_date(value)
            if entry_date:
                days = business_days_between(entry_date, today)
                entry_date_text = entry_date.isoformat()
            elif isinstance(value, (int, float)) and value >= 0:
                # Allow direct banking-day counters in cells (e.g., 2, 4, 65).
                days = int(value)
                entry_date_text = ""
            elif isinstance(value, str) and value.strip().isdigit():
                days = int(value.strip())
                entry_date_text = ""
            else:
                continue

            status = classify_status(days, settings)
            recipient = self._resolve_recipient(ws, row, settings)
            emailed = self._is_emailed(ws, row, settings)

            entries.append(
                MonitoredEntry(
                    sheet_name=ws.title,
                    row=row,
                    column=index_to_col(col_idx),
                    cell=cell_ref,
                    entry_date=entry_date_text,
                    days=days,
                    status=status,
                    recipient=recipient,
                    emailed=emailed,
                )
            )

        wb.close()
        return entries

    def mark_emailed(self, settings: AppSettings, rows: Iterable[int]) -> int:
        if not settings.email_sent_column:
            return 0

        workbook_path = Path(settings.excel_file_path)
        if not workbook_path.exists():
            return 0

        wb = load_workbook(workbook_path, read_only=False, data_only=False, keep_links=False)
        ws = wb[settings.watch.sheet_name] if settings.watch.sheet_name in wb.sheetnames else wb.active
        target_col = col_to_index(settings.email_sent_column)

        updated = 0
        for row in dedupe_rows(rows):
            cell = ws.cell(row=row, column=target_col)
            if str(cell.value).strip().lower() in {"yes", "true", "1"}:
                continue
            cell.value = "Yes"
            updated += 1

        if updated:
            wb.save(workbook_path)
        wb.close()
        return updated

    def serialize(self, entries: list[MonitoredEntry]) -> list[dict[str, object]]:
        return [asdict(e) for e in entries]

    def _resolve_recipient(self, ws, row: int, settings: AppSettings) -> str:
        cfg = settings.email
        if cfg.recipient_mode == "excel_column" and cfg.email_column:
            col_idx = col_to_index(cfg.email_column)
            val = ws.cell(row=row, column=col_idx).value
            if val:
                return str(val).strip()
        return cfg.global_recipient.strip()

    def _is_emailed(self, ws, row: int, settings: AppSettings) -> bool:
        if not settings.email_sent_column:
            return False
        col_idx = col_to_index(settings.email_sent_column)
        raw = ws.cell(row=row, column=col_idx).value
        return str(raw).strip().lower() in {"yes", "true", "1", "sent"}

    def _iter_targets(self, ws, settings: AppSettings):
        mode = settings.watch.mode

        if mode == "rows_all_columns":
            rows = [int(r) for r in parse_list_csv(settings.watch.row_list) if r.isdigit()]
            if not rows:
                rows = list(range(settings.watch.start_row, settings.watch.end_row + 1))

            # Only this mode needs worksheet width because it intentionally scans all columns.
            max_col = ws.max_column or col_to_index(settings.watch.end_col or settings.watch.start_col or "A")
            start_col = 1
            end_col = max_col
            for row in rows:
                if row < 1:
                    continue
                for col_idx in range(start_col, end_col + 1):
                    yield row, col_idx
            return

        if mode == "columns_all_rows":
            columns = [normalize_col(c) for c in parse_list_csv(settings.watch.column_list)]
            if not columns:
                columns = [normalize_col(settings.watch.start_col)]
            col_indices = [col_to_index(c) for c in columns if c]
            start_row = max(1, settings.watch.start_row)
            end_row = max(start_row, settings.watch.end_row)
            for row in range(start_row, end_row + 1):
                for col_idx in col_indices:
                    yield row, col_idx
            return

        # Default mode: rectangular range.
        start_col = col_to_index(settings.watch.start_col)
        end_col = col_to_index(settings.watch.end_col or settings.watch.start_col)
        if end_col < start_col:
            start_col, end_col = end_col, start_col
        start_row = max(1, settings.watch.start_row)
        end_row = max(start_row, settings.watch.end_row)

        for row in range(start_row, end_row + 1):
            for col_idx in range(start_col, end_col + 1):
                yield row, col_idx
